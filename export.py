import openpyxl
import os

def export_one(data, filename):

    save_dir = './csv'
    if not os.path.exists(save_dir):
        os.mkdir(save_dir)
    filepath = os.path.join(save_dir, filename)
    if os.path.exists(filepath):
        workbook = openpyxl.load_workbook(filepath)
        worksheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet['A1'] = 'Link'
        worksheet['B1'] = 'Title'
        
    start_row = worksheet.max_row + 1

    worksheet.cell(row=start_row, column=1).value = '=HYPERLINK("{}")'.format(data["link"])
    worksheet.cell(row=start_row, column=2).value = data["title"]

    # Save the workbook
    workbook.save(filepath)
    workbook.close()